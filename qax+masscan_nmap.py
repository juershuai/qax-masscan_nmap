import socket
import threading
import openpyxl
from openpyxl.styles import Alignment
import time
import re
import sys
import collections
#获取excel表信息--IP端口去重--探测端口存活状态--重新写入excel

def globals():#全局变量
    global openPortList
    global masscanNmapList
    global openPortDict
    global wsWebList
    global ip_arr
    global web_arr
    openPortList = []
    masscanNmapList = []
    openPortDict = {}
    wsWebList = []
    ip_arr = []  # masscan_nmap_ip
    web_arr = []  # masscan_nmap_web

def masscanNmap():
    ff = open('scan_url_port.txt', 'r', encoding='utf-8')
    str_replace = ff.read().replace('-', '')
    str_splot = str_replace.split('+')
    ip_str = str_splot[0].strip().split('\\n')[0].split('\n')
    web_str = str_splot[12].strip().replace(' ', '').split('\\n')[0].split('\n')
    for i in ip_str:
        ip = i.split('\t')[0].split(':')[0]
        port = i.split('\t')[0].split(':')[1]
        proto = i.split('\t')[1]
        ip_arr.append([ip,port,proto])# IP/端口/协议
    for i in web_str:
        ip = i.strip('|').split('|')[0].split('://')[1].split(':')[0]
        port = i.strip('|').split('|')[0].split('://')[1].split(':')[1]
        domain = '-'
        midd = i.strip('|').split('|')[1]
        title = i.strip('|').split('|')[2]
        web_arr.append([ip,port,midd,domain,title])# IP/端口/中间件/域名/title
    ff.close()

def threadStart():
    portScanJoin=[]#线程列表
    for i in range(2,ws.max_row+1):
        time.sleep(0.1)
        portScanThread = threading.Thread(target=portScan,args=(i,5))
        portScanThread.start()
        portScanJoin.append(portScanThread)
    for i in portScanJoin:#等待线程全部执行完成
        i.join()

def portScan(sum,time):#端口存活探测
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    sock.settimeout(time)  # 设置超时时间
    strSum = str(sum)
    result = sock.connect_ex((ws['E'+strSum].value, int(ws['I'+strSum].value)))
    if 0 == result:#开放
        openPort = [ws['E' + strSum].value,ws['H' + strSum].value,ws['I' + strSum].value,ws['J' + strSum].value,'-','-']#IP/协议/端口/中间件/域名/title
        openPortList.append(openPort)

def masscanNmapScan():#masscan扫描资产与平台资产去重
    for i in ip_arr:
        isTrue = True
        for j in openPortList:
            if (str(j[0]) == str(i[0])) and (str(j[2]) == str(i[1])):
                isTrue = False
                break
        if isTrue:
            webArr = findWebArr(i)
            masscanNmapList.append(webArr)

def findWebArr(i):#masscan中查找对应ip的web信息
    for j in web_arr:
        if (j[0] == i[0]) and (j[1]==i[1]):
            return [i[0],i[2],i[1],j[2],j[3],j[4]]
        else:
            return [i[0],i[2],i[1],'','','']

def remove():#开放端口去重(域名和IP同时扫描会有重复)
    removeOpenPort = [list(j) for j in set(tuple(i) for i in openPortList)]
    return removeOpenPort

def listToDict():#ip作为键写入字典(字典格式方便后续写入excel)
    for i in removeOpenPortList:
        if i[0] not in openPortDict:
            openPortDict.setdefault(i[0], [i])#包括ip字段所有数据写进字典
        else:
            openPortDict[i[0]] += [i]#包括ip字段所有数据写进字典

def searchDomainTitle():#去'web资产'中查找域名和title信息
    ws_web = wb['web资产']
    for i in range(2, ws_web.max_row):
        strI = str(i)
        wsWebList.append([ws_web['E'+strI].value,ws_web['F'+strI].value,ws_web['I'+strI].value,ws_web['C'+strI].value])#IP/域名/端口/title
    for i in openPortDict.values():
       for j in i:
           j[4],j[5]=findDomainTitle(j[0],j[2])#接收为str  \n换行分隔

def findDomainTitle(ip,port):#返回去重后的域名和title信息
    domainList = []
    titleList = []
    p = re.compile(
        '^(1\d{2}|2[0-4]\d|25[0-5]|[1-9]\d|[1-9])\.(1\d{2}|2[0-4]\d|25[0-5]|[1-9]\d|\d)\.(1\d{2}|2[0-4]\d|25[0-5]|[1-9]\d|\d)\.(1\d{2}|2[0-4]\d|25[0-5]|[1-9]\d|\d)$')
    for i in wsWebList:
        if (ip == i[0]) & (port == i[2]):
            if (i[1] not in domainList) & (p.match(i[1]) == None):#正则过滤域名为ip
                domainList.append(i[1])
            if (i[3] not in titleList) & (i[3] != None) & (i[3] != '-'):
                titleList.append(i[3])
    domainStr = ''
    titleStr = ''
    if len(domainList) > 0:
        if len(domainList) == 1:
            domainStr = domainList[0]
        else:
            for i in domainList:
                domainStr = domainStr + i + '\r\n'
    if len(titleList) > 0:
        if len(titleList) == 1:
            titleStr = titleList[0]
        else:
            for i in titleList:
                titleStr = titleStr + i + '\r\n'
    return domainStr,titleStr

def saveExcel():
    first = ['IP','协议','端口','服务','域名','title']
    assert_wb = openpyxl.Workbook()
    assert_ws = assert_wb.active
    assert_ws.title='服务器资产'
    assert_ws.append(str(fir) for fir in first)
    newPortDict = collections.OrderedDict(sorted(openPortDict.items()))#字典按key排序
    for i in newPortDict.values():
        for j in i:
            assert_ws.append(str(fir) for fir in j)
    index = 1
    for i in newPortDict.keys():
        assert_ws.merge_cells('A'+str(index+1)+':A'+str(len(newPortDict[i])+index))
        index+=len(newPortDict[i])
    center = Alignment(horizontal='left', vertical='center', wrap_text=False)#第一列居中
    for i in assert_ws['A2:A'+str(assert_ws.max_row)]:
        for j in i:
            j.alignment = center

    assert_wb.create_sheet('子域名资产')#子域名资产
    assert_ws2 = assert_wb['子域名资产']
    for i,row in enumerate(wb['子域名资产'].iter_rows()):
        for j,cell in enumerate(row):
            assert_ws2.cell(row=i+1, column=j+1, value=cell.value)
    assert_wb.save(sys.path[0]+'\互联网资产发现清单.xlsx')

if __name__ == "__main__":
    print("格式：python3 run.py 资产发现_3.xlsx")
    start = time.time()
    wb = openpyxl.load_workbook(sys.argv[1])
    ws = wb['服务器资产']
    globals()
    masscanNmap()
    threadStart()
    masscanNmapScan()
    openPortList += masscanNmapList
    removeOpenPortList = remove()
    listToDict()
    searchDomainTitle()
    saveExcel()
    end = time.time()
    dtime = end - start
    print("程序运行时间：%.4s s" % dtime)
